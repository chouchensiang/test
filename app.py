#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
app.py

Flask 後端服務，提供化學品相容性查詢 API。

用法：
  python app.py

API 端點：
  GET  /                     - 首頁（查詢介面）
  POST /api/query            - 查詢相容性
  POST /api/matrix           - 上傳 Excel 產生矩陣
  GET  /api/stats            - 取得統計資料
"""

import os
import sys
import sqlite3
import webbrowser
import threading
from io import BytesIO
from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import pandas as pd

# 判斷是否為打包後的 EXE
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__, static_folder=BASE_DIR, static_url_path='')
CORS(app)

# === 設定 ===
REPORT_DIR_CANDIDATES = ["report", "reports"]
DB_FILENAME = "ChemCompatibility.db"
DB_PATH = None

for report_dir in REPORT_DIR_CANDIDATES:
    candidate = os.path.join(BASE_DIR, report_dir, DB_FILENAME)
    if os.path.exists(candidate):
        DB_PATH = candidate
        break

if DB_PATH is None:
    DB_PATH = os.path.join(BASE_DIR, REPORT_DIR_CANDIDATES[0], DB_FILENAME)

UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def get_db_connection():
    """取得資料庫連線"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def normalize_compatibility(value):
    """將資料庫中的相容性值正規化為 Y/N/C/X。"""
    if value is None:
        return None

    if isinstance(value, (int, float)):
        if int(value) == 1:
            return 'Y'
        if int(value) == 0:
            return 'N'

    value_str = str(value).strip().upper()
    if value_str in {'1', 'Y', 'YES', 'TRUE'}:
        return 'Y'
    if value_str in {'0', 'N', 'NO', 'FALSE'}:
        return 'N'
    if value_str in {'C', 'X'}:
        return value_str
    return value_str or None


def get_compatible_filter_values(filter_value):
    """回傳查詢條件可接受的資料庫值。"""
    mapping = {
        'Y': ['Y', '1'],
        'N': ['N', '0'],
        'C': ['C'],
        'X': ['X'],
    }
    return mapping.get(filter_value, [])


@app.route('/')
def index():
    """首頁"""
    return send_from_directory(BASE_DIR, 'web.html')


@app.route('/api/query', methods=['POST'])
def query_compatibility():
    """查詢化學品相容性"""
    try:
        data = request.get_json() or {}
        cas_a = data.get('cas_a', '').strip()
        cas_b = data.get('cas_b', '').strip()
        compatible_filter = data.get('compatible_filter', '').strip().upper()
        limit = min(int(data.get('limit', 100)), 1000)
        
        # 必須至少輸入一個查詢條件
        if not cas_a and not cas_b and not compatible_filter:
            return jsonify({
                "success": False,
                "error": "請至少輸入一個查詢條件（CAS No. 或相容性篩選）"
            }), 400
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        conditions = []
        params = []
        
        if cas_a:
            conditions.append("(Chemical_A = ? OR Chemical_B = ?)")
            params.extend([cas_a, cas_a])
        
        if cas_b:
            conditions.append("(Chemical_A = ? OR Chemical_B = ?)")
            params.extend([cas_b, cas_b])
        
        if compatible_filter in ['Y', 'N', 'C', 'X']:
            filter_values = get_compatible_filter_values(compatible_filter)
            placeholders = ','.join(['?' for _ in filter_values])
            conditions.append(f"UPPER(CAST(Compatible AS TEXT)) IN ({placeholders})")
            params.extend(filter_values)
        
        sql = "SELECT Chemical_A, Chemical_B, Compatible FROM compatibility"
        if conditions:
            sql += " WHERE " + " AND ".join(conditions)
        sql += f" LIMIT {limit}"
        
        cur.execute(sql, params)
        rows = cur.fetchall()
        conn.close()
        
        results = []
        for row in rows:
            result = dict(row)
            result['Compatible'] = normalize_compatibility(row['Compatible'])
            results.append(result)
        
        return jsonify({
            "success": True,
            "count": len(results),
            "data": results
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/check', methods=['POST'])
def check_pair():
    """檢查兩個化學品的相容性"""
    try:
        data = request.get_json() or {}
        cas_a = data.get('cas_a', '').strip()
        cas_b = data.get('cas_b', '').strip()
        
        if not cas_a or not cas_b:
            return jsonify({"success": False, "error": "請提供 cas_a 和 cas_b"}), 400
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        sql = """
        SELECT Compatible FROM compatibility 
        WHERE (Chemical_A = ? AND Chemical_B = ?) 
           OR (Chemical_A = ? AND Chemical_B = ?)
        LIMIT 1
        """
        cur.execute(sql, (cas_a, cas_b, cas_b, cas_a))
        row = cur.fetchone()
        conn.close()
        
        if row:
            return jsonify({
                "success": True,
                "cas_a": cas_a,
                "cas_b": cas_b,
                "compatible": normalize_compatibility(row["Compatible"]),
                "found": True
            })
        else:
            return jsonify({
                "success": True,
                "cas_a": cas_a,
                "cas_b": cas_b,
                "compatible": None,
                "found": False,
                "message": "找不到此配對的相容性資料"
            })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/matrix', methods=['POST'])
def generate_matrix():
    """
    上傳 Excel/CSV，產生料號相容性矩陣
    
    檔案格式：
      - 第一欄：料號（如 A, B, C...）
      - 第二欄：CAS No.（多個用換行分隔）
    
    Response JSON:
      {
        "success": true,
        "items": ["A", "B", "C", ...],
        "matrix": [
          ["", "A", "B", "C", ...],
          ["A", "-", "Y", "N", ...],
          ...
        ],
        "details": {
          "A|B": {"result": "Y", "pairs": [...]},
          ...
        },
        "summary": {"Y": 10, "N": 5, "NA": 3}
      }
    """
    try:
        if 'file' not in request.files:
            return jsonify({"success": False, "error": "請上傳檔案"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"success": False, "error": "請選擇檔案"}), 400
        
        filename = file.filename.lower()
        
        # 讀取檔案
        try:
            if filename.endswith('.csv'):
                df = pd.read_csv(file, encoding='utf-8')
            else:
                df = pd.read_excel(file, engine='openpyxl')
        except Exception as e:
            return jsonify({"success": False, "error": f"無法讀取檔案: {e}"}), 400
        
        if len(df.columns) < 2:
            return jsonify({"success": False, "error": "檔案需要至少兩欄（料號、CAS No.）"}), 400
        
        # 智慧識別欄位名稱
        item_col = None
        cas_col = None
        
        for col in df.columns:
            col_upper = col.upper().replace(' ', '').replace('_', '').replace('.', '')
            # 識別料號欄位
            if col_upper in ['料號', '奇美名稱', 'ITEMNO', 'ITEM', 'PARTNO', 'PART', 'PRODUCTNO', 'PRODUCT', 'CODE', 'NO', '品號', '產品編號', '名稱']:
                item_col = col
            # 識別 CAS 欄位
            elif col_upper in ['CAS', 'CASNO', 'CASNUMBER', 'CASNUM']:
                cas_col = col
        
        # 如果沒找到，使用預設（第一欄=料號，第二欄=CAS）
        if item_col is None:
            item_col = df.columns[0]
        if cas_col is None:
            # 找第一個不是 item_col 的欄位
            for col in df.columns:
                if col != item_col:
                    cas_col = col
                    break
        
        # 解析每個料號的 CAS 清單
        item_cas_map = {}  # {料號: [CAS1, CAS2, ...]}
        all_cas_set = set()
        
        for _, row in df.iterrows():
            item = str(row[item_col]).strip()
            cas_str = str(row[cas_col]).strip()
            
            if not item or item.lower() == 'nan':
                continue
            
            # 解析多個 CAS（可能用換行、逗號、分號分隔）
            cas_list = []
            for sep in ['\n', ',', ';', '|']:
                if sep in cas_str:
                    cas_list = [c.strip() for c in cas_str.split(sep) if c.strip() and c.strip().lower() != 'nan']
                    break
            
            if not cas_list:
                cas_list = [cas_str] if cas_str and cas_str.lower() != 'nan' else []
            
            if cas_list:
                item_cas_map[item] = cas_list
                all_cas_set.update(cas_list)
        
        if len(item_cas_map) == 0:
            return jsonify({"success": False, "error": "找不到有效的料號資料"}), 400
        
        # 查詢資料庫建立相容性對照表
        conn = get_db_connection()
        cur = conn.cursor()
        
        compat_map = {}  # (cas_a, cas_b) -> Compatible
        all_cas_list = list(all_cas_set)
        
        # 分批查詢（避免 SQL 太長）
        batch_size = 100
        for i in range(0, len(all_cas_list), batch_size):
            batch = all_cas_list[i:i+batch_size]
            placeholders = ','.join(['?' for _ in batch])
            sql = f"""
            SELECT Chemical_A, Chemical_B, Compatible 
            FROM compatibility 
            WHERE Chemical_A IN ({placeholders}) OR Chemical_B IN ({placeholders})
            """
            cur.execute(sql, batch + batch)
            
            for row in cur.fetchall():
                a = row['Chemical_A']
                b = row['Chemical_B']
                c = normalize_compatibility(row['Compatible'])
                if a in all_cas_set and b in all_cas_set:
                    compat_map[(a, b)] = c
                    compat_map[(b, a)] = c
        
        conn.close()
        
        # 建立料號清單
        items = list(item_cas_map.keys())
        n = len(items)
        
        # 建立矩陣和詳細資訊
        matrix = []
        details = {}  # 詳細比對結果
        summary = {"Y": 0, "N": 0, "NA": 0}
        
        # 表頭
        header = [""] + items
        matrix.append(header)
        
        # 判斷兩個料號的相容性
        def check_item_compatibility(item_a, item_b):
            """
            比對兩個料號的相容性
            規則：
            - 任一配對為 N → 整體 N
            - 全部為 Y → 整體 Y
            - 有任何配對不在資料庫 → NA
            """
            cas_a_list = item_cas_map[item_a]
            cas_b_list = item_cas_map[item_b]
            
            pairs = []
            has_n = False
            has_na = False
            all_y = True
            
            for cas_a in cas_a_list:
                for cas_b in cas_b_list:
                    if cas_a == cas_b:
                        continue  # 同一個 CAS 跳過
                    
                    compat = compat_map.get((cas_a, cas_b), None)
                    pair_result = compat if compat is not None else "NA"
                    pair_info = {
                        "cas_a": cas_a,
                        "cas_b": cas_b,
                        "result": pair_result
                    }
                    pairs.append(pair_info)
                    
                    if compat == 'N':
                        has_n = True
                        all_y = False
                    elif compat == 'Y':
                        pass  # 繼續
                    else:
                        has_na = True
                        all_y = False
            
            # 判斷最終結果
            if has_n:
                result = "N"
            elif all_y and len(pairs) > 0:
                result = "Y"
            else:
                result = "NA"
            
            return result, pairs
        
        # 每一列（只顯示下三角，上三角反白）
        for i, item_a in enumerate(items):
            row = [item_a]
            for j, item_b in enumerate(items):
                if i == j:
                    row.append("-")  # 對角線
                elif i < j:
                    row.append("")  # 上三角反白
                else:
                    # 下三角：計算相容性
                    result, pairs = check_item_compatibility(item_a, item_b)
                    row.append(result)
                    
                    # 儲存詳細資訊
                    key = f"{item_a}|{item_b}"
                    details[key] = {
                        "result": result,
                        "item_a": item_a,
                        "item_b": item_b,
                        "cas_a": item_cas_map[item_a],
                        "cas_b": item_cas_map[item_b],
                        "pairs": pairs
                    }
                    summary[result] = summary.get(result, 0) + 1
            
            matrix.append(row)
        
        # 網頁顯示用的矩陣（最多50筆）
        display_limit = 50
        if len(items) > display_limit:
            display_items = items[:display_limit]
            display_matrix = [matrix[0][:display_limit+1]]  # 表頭
            for i in range(1, display_limit+1):
                display_matrix.append(matrix[i][:display_limit+1])
        else:
            display_items = items
            display_matrix = matrix
        
        return jsonify({
            "success": True,
            "items": display_items,
            "all_items": items,  # 完整清單（用於下載）
            "count": len(items),
            "display_count": len(display_items),
            "matrix": display_matrix,
            "full_matrix": matrix,  # 完整矩陣（用於下載）
            "details": details,
            "item_cas_map": item_cas_map,
            "summary": summary,
            "truncated": len(items) > display_limit
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/matrix/download', methods=['POST'])
def download_matrix():
    """下載完整矩陣為 Excel"""
    try:
        data = request.get_json() or {}
        matrix = data.get('full_matrix') or data.get('matrix', [])
        
        if not matrix:
            return jsonify({"success": False, "error": "沒有矩陣資料"}), 400
        
        # 轉成 DataFrame
        df = pd.DataFrame(matrix[1:], columns=matrix[0])
        
        # 寫入 Excel（含顏色）
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        COLOR_MAP = {
            'Y':  ('C3E6CB', '155724'),  # 綠底深綠字
            'N':  ('F5C6CB', '721C24'),  # 紅底深紅字
            'C':  ('FFEEBA', '856404'),  # 黃底
            'X':  ('D6D8DB', '383D41'),  # 灰底
            'NA': ('F8F9FA', '999999'),  # 淡灰底
            '-':  ('E9ECEF', '999999'),  # 對角線
            '':   ('F8F9FA', '000000'),  # 上三角空白
        }

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Compatibility Matrix')
            ws = writer.sheets['Compatibility Matrix']

            # 自動調整欄寬（最小 12、最大 28，料號列最小 18）
            for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
                col_letter = get_column_letter(col_idx)
                col_data = [str(cell.value) if cell.value is not None else '' for cell in col_cells]
                
                # 計算最大內容長度
                max_length = max([len(str(cell)) for cell in col_data]) if col_data else 10
                
                # 第一欄（料號）最少 18，其他欄 12-28
                if col_idx == 1:
                    width = max(18, min(max_length + 2, 28))
                else:
                    width = max(12, min(max_length + 2, 28))
                
                ws.column_dimensions[col_letter].width = width

            # 套用顏色（跳過第一列表頭，從第 2 列開始）
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                for col_idx, cell in enumerate(row):
                    val = str(cell.value) if cell.value is not None else ''
                    if val in COLOR_MAP:
                        bg, fg = COLOR_MAP[val]
                        cell.fill = PatternFill(fill_type='solid', fgColor=bg)
                        cell.font = Font(bold=(val in ('Y', 'N')), color=fg)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # 表頭格式
            for cell in ws[1]:
                cell.fill = PatternFill(fill_type='solid', fgColor='E9ECEF')
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 凍結首列首欄（在 B2 位置凍結）
            ws.freeze_panes = 'B2'

        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='compatibility_matrix.xlsx'
        )
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/stats', methods=['GET'])
def get_stats():
    """取得資料庫統計資訊"""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute("SELECT COUNT(*) FROM compatibility")
        total = cur.fetchone()[0]
        
        cur.execute("""
            SELECT Compatible, COUNT(*) as cnt 
            FROM compatibility 
            GROUP BY Compatible 
            ORDER BY cnt DESC
        """)
        distribution = {}
        for row in cur.fetchall():
            normalized = normalize_compatibility(row["Compatible"])
            distribution[normalized] = distribution.get(normalized, 0) + row["cnt"]
        
        cur.execute("SELECT COUNT(DISTINCT Chemical_A) FROM compatibility")
        unique_a = cur.fetchone()[0]
        
        conn.close()
        
        return jsonify({
            "success": True,
            "total_pairs": total,
            "unique_chemicals": unique_a,
            "distribution": distribution
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/shutdown', methods=['POST'])
def shutdown():
    """關閉伺服器"""
    print("[INFO] 收到關閉請求，正在結束程式...")
    # 給一點時間讓回應送出
    def shutdown_server():
        import time
        time.sleep(0.5)
        os._exit(0)
    threading.Thread(target=shutdown_server, daemon=True).start()
    return jsonify({"success": True, "message": "伺服器正在關閉"})


def open_browser():
    """延遲開啟瀏覽器"""
    import time
    time.sleep(1.5)
    webbrowser.open('http://127.0.0.1:5000')


if __name__ == '__main__':
    print("=" * 60)
    print("[SERVER] 化學品相容性查詢系統")
    print("=" * 60)
    print(f"[INFO] Database: {DB_PATH}")
    print(f"[INFO] Server: http://127.0.0.1:5000")
    print("=" * 60)
    
    # 只在主進程開啟瀏覽器（避免 reloader 重複開啟）
    import os
    if os.environ.get('WERKZEUG_RUN_MAIN') != 'true' and not getattr(sys, 'frozen', False):
        # 開發模式：等 reloader 啟動後才開
        pass
    elif os.environ.get('WERKZEUG_RUN_MAIN') == 'true' or getattr(sys, 'frozen', False):
        # reloader 子進程或 EXE 模式：開啟瀏覽器
        threading.Thread(target=open_browser, daemon=True).start()
    
    # 如果是 EXE 模式，不顯示 debug 訊息
    if getattr(sys, 'frozen', False):
        app.run(host='127.0.0.1', port=5000, debug=False, use_reloader=False)
    else:
        app.run(host='127.0.0.1', port=5000, debug=True, use_reloader=True)
